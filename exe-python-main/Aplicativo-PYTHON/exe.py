import customtkinter as ctk
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from tkinter import messagebox
from datetime import datetime
import pandas as pd

# Variável global para armazenar a data do último cálculo
ultima_data_calculo = None

# Função para formatar a data automaticamente enquanto o usuário digita
def atualizar_data(event):
    entrada_data = entry_data.get().replace("/", "")  # Remove barras para reformatar

    if len(entrada_data) > 8:
        entrada_data = entrada_data[:8]
    if len(entrada_data) > 2:
        entrada_data = entrada_data[:2] + '/' + entrada_data[2:]  # Adiciona a primeira barra
    if len(entrada_data) > 5:
        entrada_data = entrada_data[:5] + '/' + entrada_data[5:]  # Adiciona a segunda barra
    
    entry_data.delete(0, 'end')
    entry_data.insert(0, entrada_data)

# Função para calcular e salvar na planilha
def calcular_e_salvar():
    nome_produto = entry_nome.get()
    valor_compra = entry_valor_compra.get()
    valor_venda = entry_valor_venda.get()
    quantidade = entry_quantidade.get()
    observacao = entry_observacao.get("1.0", "end").strip()
    data_entrada = entry_data.get()

    if not nome_produto or not valor_compra or not valor_venda or not quantidade or not data_entrada:
        messagebox.showerror("Erro", "Por favor, preencha todos os campos obrigatórios.")
        return

    try:
        # Converter valores para float e int
        valor_compra = float(valor_compra)
        valor_venda = float(valor_venda)
        quantidade = int(quantidade)

        # Cálculo de lucro, faturamento e gasto total
        faturamento = valor_venda * quantidade
        lucro = (valor_venda - valor_compra) * quantidade
        gasto_total = valor_compra * quantidade

        file_name = "gestao_produtos.xlsx"

        # Tenta carregar uma planilha existente ou criar uma nova
        try:
            workbook = load_workbook(file_name)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Relatório de Produtos"
            headers = ["Nome do Produto", "Data da Entrada", "Quantidade", "Valor de Compra", "Valor de Venda", "Gasto Total", "Faturamento", "Lucro", "Observação"]
            sheet.append(headers)

        # Adiciona dados
        data = [nome_produto, data_entrada, quantidade, valor_compra, valor_venda, gasto_total, faturamento, lucro, observacao]
        sheet.append(data)

        # Formatação dos cabeçalhos
        if sheet.max_row == 2:  # Se é um novo arquivo
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Ajuste das larguras das colunas
            column_widths = [max(len(str(item)) for item in column) for column in zip(*sheet.iter_rows(values_only=True))]
            for i, column_width in enumerate(column_widths, start=1):
                column_letter = chr(64 + i)
                sheet.column_dimensions[column_letter].width = column_width + 2

        workbook.save(file_name)
        messagebox.showinfo("Sucesso", "Dados salvos com sucesso na planilha gestao_produtos.xlsx")

        # Limpar todos os campos
        entry_nome.delete(0, 'end')
        entry_quantidade.delete(0, 'end')
        entry_valor_compra.delete(0, 'end')
        entry_valor_venda.delete(0, 'end')
        entry_observacao.delete("1.0", 'end')
        entry_data.delete(0, 'end')

    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro ao salvar a planilha: {str(e)}")

# Função para calcular e exibir os totais mensais (apenas para os dados novos)
def calcular_totais_mensais():
    global ultima_data_calculo
    file_name = "gestao_produtos.xlsx"
    try:
        # Carregar os dados da planilha
        df = pd.read_excel(file_name, sheet_name="Relatório de Produtos")

        # Garantir que a coluna de data esteja no formato correto
        df['Data da Entrada'] = pd.to_datetime(df['Data da Entrada'], format='%d/%m/%Y', errors='coerce')

        # Log para verificar se as datas foram lidas corretamente
        print("Datas na planilha (convertidas para datetime):")
        print(df['Data da Entrada'])

        if df['Data da Entrada'].isnull().all():
            messagebox.showerror("Erro", "Não há datas válidas na planilha.")
            return

        # Filtrar os dados a partir da última data de cálculo
        if ultima_data_calculo is not None:
            df_filtrado = df[df['Data da Entrada'] > ultima_data_calculo]
        else:
            df_filtrado = df  # Se nunca foi feito um cálculo, pega todos os dados

        # Log para ver como o filtro está funcionando
        print("Dados filtrados (após última data de cálculo):")
        print(df_filtrado)

        if df_filtrado.empty:
            messagebox.showinfo("Sem Dados", "Não há registros novos para o cálculo.")
            return

        # Obter o mês e ano atual
        data_atual = datetime.now()
        mes_atual = data_atual.month
        ano_atual = data_atual.year

        # Log para verificar o mês e ano atual
        print(f"Mês atual: {mes_atual}, Ano atual: {ano_atual}")

        # Filtrar os dados do mês e ano atual
        df_filtrado = df_filtrado[(df_filtrado['Data da Entrada'].dt.month == mes_atual) & (df_filtrado['Data da Entrada'].dt.year == ano_atual)]

        # Log para verificar o filtro do mês e ano atual
        print("Dados filtrados para o mês atual:")
        print(df_filtrado)

        if df_filtrado.empty:
            messagebox.showinfo("Sem Dados", "Não há registros para o mês atual.")
            return

        # Calcular totais de lucro, faturamento e gasto total
        totais = df_filtrado[['Lucro', 'Faturamento', 'Gasto Total']].sum()
        lucro_total = totais['Lucro']
        faturamento_total = totais['Faturamento']
        gasto_total = totais['Gasto Total']

        # Exibir resultados
        label_resultado.configure(text=f"Lucro Total: R${lucro_total:.2f}\nFaturamento Total: R${faturamento_total:.2f}\nGasto Total: R${gasto_total:.2f}")

        # Atualizar a planilha com os totais mensais
        workbook = load_workbook(file_name)
        sheet = workbook.active

        # Encontrar a última linha da planilha
        ultima_linha = sheet.max_row + 1

        # Adicionar os totais na nova linha abaixo dos dados
        sheet.cell(row=ultima_linha, column=1, value="Totais Mensais")
        sheet.cell(row=ultima_linha, column=6, value=gasto_total)  # Gasto Total
        sheet.cell(row=ultima_linha, column=7, value=faturamento_total)  # Faturamento Total
        sheet.cell(row=ultima_linha, column=8, value=lucro_total)  # Lucro Total

        workbook.save(file_name)

        # Atualizar a data do último cálculo
        ultima_data_calculo = datetime.now()

    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro ao calcular os totais mensais: {str(e)}")

# Função para alternar o modo de aparência
def alternar_tema():
    global tema_atual
    if tema_atual == "dark":
        ctk.set_appearance_mode("light")
        btn_alternar_tema.configure(text="Modo Escuro")
        tema_atual = "light"
    else:
        ctk.set_appearance_mode("dark")
        btn_alternar_tema.configure(text="Modo Claro")
        tema_atual = "dark"

# Configurando a interface gráfica com CustomTkinter
ctk.set_appearance_mode("dark")  # Modo escuro inicial
ctk.set_default_color_theme("blue")  # Tema padrão

tema_atual = "dark"

root = ctk.CTk()  # Janela principal
root.title("Gestão de Produtos LK")
root.geometry("800x600")  # Tamanho fixo da janela
root.resizable(False, False)
root.grid_rowconfigure(0, weight=1)  # Centraliza verticalmente
root.grid_columnconfigure(0, weight=1)  # Centraliza horizontalmente

# Frame principal para centralização
frame = ctk.CTkFrame(root)
frame.grid(row=0, column=0, padx=20, pady=20, sticky="nsew")

# Configurando os pesos das linhas e colunas para expandir e preencher
frame.grid_rowconfigure((0, 1, 2, 3, 4, 5, 6, 7, 8, 9), weight=1)
frame.grid_columnconfigure((0, 1), weight=1)

# Nome do produto
ctk.CTkLabel(frame, text="Nome do Produto:").grid(row=0, column=0, padx=10, pady=10, sticky="e")
entry_nome = ctk.CTkEntry(frame, width=200, placeholder_text="Digite o nome")
entry_nome.grid(row=0, column=1, padx=10, pady=10, sticky="w")

# Data da Entrada
ctk.CTkLabel(frame, text="Data da Venda:").grid(row=1, column=0, padx=10, pady=10, sticky="e")
entry_data = ctk.CTkEntry(frame, width=200, placeholder_text="Digite a data (DDMMYY)")
entry_data.grid(row=1, column=1, padx=10, pady=10, sticky="w")
entry_data.bind("<KeyRelease>", atualizar_data)  # Formata a data enquanto digita

# Quantidade
ctk.CTkLabel(frame, text="Quantidade:").grid(row=2, column=0, padx=10, pady=10, sticky="e")
entry_quantidade = ctk.CTkEntry(frame, width=200, placeholder_text="Digite a quantidade")
entry_quantidade.grid(row=2, column=1, padx=10, pady=10, sticky="w")

# Valor de Compra
ctk.CTkLabel(frame, text="Valor de Compra:").grid(row=3, column=0, padx=10, pady=10, sticky="e")
entry_valor_compra = ctk.CTkEntry(frame, width=200, placeholder_text="Digite o valor de compra")
entry_valor_compra.grid(row=3, column=1, padx=10, pady=10, sticky="w")

# Valor de Venda
ctk.CTkLabel(frame, text="Valor de Venda:").grid(row=4, column=0, padx=10, pady=10, sticky="e")
entry_valor_venda = ctk.CTkEntry(frame, width=200, placeholder_text="Digite o valor de venda")
entry_valor_venda.grid(row=4, column=1, padx=10, pady=10, sticky="w")

# Observação
ctk.CTkLabel(frame, text="Observação:").grid(row=5, column=0, padx=10, pady=10, sticky="ne")
entry_observacao = ctk.CTkTextbox(frame, width=200, height=100)
entry_observacao.grid(row=5, column=1, padx=10, pady=10, sticky="w")

# Botão Calcular
btn_calcular = ctk.CTkButton(frame, text="Calcular e Salvar", command=calcular_e_salvar)
btn_calcular.grid(row=6, column=0, columnspan=2, padx=10, pady=10)

# Botão Calcular Totais
btn_calcular_totais = ctk.CTkButton(frame, text="Calcular Totais Mensais", command=calcular_totais_mensais)
btn_calcular_totais.grid(row=7, column=0, columnspan=2, padx=10, pady=10)

# Resultado
label_resultado = ctk.CTkLabel(frame, text="Lucro Total: R$0.00\nFaturamento Total: R$0.00")
label_resultado.grid(row=8, column=0, columnspan=2, padx=10, pady=10)

# Botão Alternar Tema
btn_alternar_tema = ctk.CTkButton(frame, text="Modo Claro", command=alternar_tema)
btn_alternar_tema.grid(row=9, column=0, columnspan=2, padx=10, pady=10)

root.mainloop()